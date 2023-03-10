

import re
import openpyxl

from src.MysqlTool import MysqlTool
from src.utils import process_symptoms


sqltool = MysqlTool()


def process(table):
    for i in range(2, table.max_row+1):
        print("--------------------------------------------------")
        # 解析字段
        sytech_name = table[i][0].value.strip()
        disease_name = table[i][1].value.strip()
        syndrome_name = table[i][2].value.strip()
        symptom_str = table[i][3].value.strip()
        detail = table[i][4].value.strip()
        operation = table[i][5].value.strip()

        # 添加症状
        symptom_list = process_symptoms(symptom_str)
        for symptom in symptom_list:
            status = 1
            cate = 1
            content = symptom
            note = symptom
            sqltool.insertDictSymptom(status, cate, content, note)

        # 添加疾病
        status = 1
        value = None
        content = disease_name
        note = disease_name
        sqltool.insertDictDisease(status, value, content, note)

        # 添加辨证
        status = 1
        symptoms = "，".join(symptom_list)
        sqltool.insertScySyndrome(status, disease_name, syndrome_name, symptoms)

        # 添加方案
        status = 1
        sqltool.insertSchSytech(status, disease_name, syndrome_name, sytech_name, detail, operation)


def process_xunzheng(table):
    for i in range(2, table.max_row+1):
        print("--------------------------------------------------")
        # 解析字段
        sytech_name = table[i][0].value.strip()
        disease_name = table[i][1].value.strip()
        syndrome_name = table[i][2].value.strip()
        symptom_str = table[i][3].value.strip()
        component = table[i][4].value.strip()
        operation = table[i][5].value.strip()
        fangan_name = syndrome_name

        # 添加症状
        symptom_list = process_symptoms(symptom_str)
        for symptom in symptom_list:
            status = 1
            cate = 1
            content = symptom
            note = symptom
            sqltool.insertDictSymptom(status, cate, content, note)

        # 添加疾病
        status = 1
        value = None
        content = disease_name
        note = disease_name
        sqltool.insertDictDisease(status, value, content, note)

        # 添加辨证
        status = 1
        symptoms = "，".join(symptom_list)
        sqltool.insertScySyndrome(status, disease_name, syndrome_name, symptoms)

        # 添加方案
        status = 1
        sqltool.insertSchXieding(status, disease_name, syndrome_name, fangan_name, component)



if __name__ == "__main__":

    sqltool.open()

    # 加载excel
    path = '/Users/huhu/糊糊工作/脉之语/工作相关/项目相关/一真堂/资料文档/规范文档/妇科儿科中医适宜技术和优势病种 (细分版).xlsx'
    data = openpyxl.load_workbook(path)

    # 处理妇科数据
    table = data["妇科"]
    process(table)

    # 处理儿科数据
    table = data["儿科"]
    process(table)

    # 处理妇科熏蒸数据
    table = data["妇科熏蒸"]
    process_xunzheng(table)

    # 处理儿科熏蒸数据
    table = data["儿科熏蒸"]
    process_xunzheng(table)

    sqltool.close()
