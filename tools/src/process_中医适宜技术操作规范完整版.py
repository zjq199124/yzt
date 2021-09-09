

import re
import openpyxl

from src.MysqlTool import MysqlTool


sqltool = MysqlTool()


def process_sytech(table):
    for i in range(2, table.max_row+1):
        id = i-1
        name = table[i][0].value.strip().replace(" ","")
        status = 1
        descrip = name
        url = ""
        if id is not None:
            sqltool.insertTsSytech(id, status, name, url, descrip)


def process_standard(table):
    id = 0
    for i in range(2, table.max_row+1):
        tid = i-1

        id += 1
        sort = 1
        name = "评估"
        content = table[i][1].value
        sqltool.insertTsStandard(id, tid, sort, name, content)

        id += 1
        sort = 1
        name = "用物准备"
        content = table[i][2].value
        sqltool.insertTsStandard(id, tid, sort, name, content)

        id += 1
        sort = 1
        name = "操作步骤"
        content = table[i][3].value
        sqltool.insertTsStandard(id, tid, sort, name, content)

        id += 1
        sort = 1
        name = "基本操作方法"
        content = table[i][4].value
        sqltool.insertTsStandard(id, tid, sort, name, content)

        id += 1
        sort = 1
        name = "禁忌症"
        content = table[i][5].value
        sqltool.insertTsStandard(id, tid, sort, name, content)

        id += 1
        sort = 1
        name = "注意事项"
        content = table[i][6].value
        sqltool.insertTsStandard(id, tid, sort, name, content)



if __name__ == "__main__":

    sqltool.open()

    # 加载excel
    path = '/Users/huhu/糊糊工作/脉之语/工作相关/项目相关/一真堂/资料文档/规范文档/中医适宜技术操作规范完整版.xlsx'
    data = openpyxl.load_workbook(path)

    # 处理数据
    table = data["Sheet1"]
    # process_sytech(table)
    process_standard(table)

    sqltool.close()
