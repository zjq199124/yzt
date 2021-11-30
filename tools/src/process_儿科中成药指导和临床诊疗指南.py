

import re
import openpyxl

from src.MysqlTool import MysqlTool
from src.utils import process_symptoms


sqltool = MysqlTool()


def process_zhongyao(table):
    for i in range(2, table.max_row+1):
        print("--------------------------------------------------")
        # 解析字段
        disease_name = table[i][0].value.strip()
        symptom_str = table[i][3].value.strip()
        syndrome_name = table[i][2].value.strip()
        name = table[i][5].value.strip().replace("（", "(").replace("）", ")")
        component = table[i][6].value.strip().replace("、", "，")
        component = component.replace("等", "").replace("。", "")

        # 添加症状
        symptom_list = process_symptoms(symptom_str)
        for symptom in symptom_list:
            status = 1
            cate = 1
            content = symptom
            note = symptom
            sqltool.insertDictSymptom(status, cate, content, note)

        # 添加疾病
        status = 1
        value = None
        content = disease_name
        note = disease_name
        sqltool.insertDictDisease(status, value, content, note)

        # 添加辨证
        status = 1
        symptoms = "，".join(symptom_list)
        sqltool.insertScySyndrome(status, disease_name, syndrome_name, symptoms)

        # 添加方案
        status = 1
        sqltool.insertSchZhongyao(status, disease_name, syndrome_name, name, component)


def process_chengyao(table):
    for i in range(2, table.max_row+1):
        print("--------------------------------------------------")
        # 解析字段
        disease_name = table[i][0].value.strip()
        name = table[i][1].value.strip()
        # component = table[i][2].value.strip()
        syndrome_name = table[i][4].value.strip()

        # 添加疾病
        status = 1
        value = None
        content = disease_name
        note = disease_name
        sqltool.insertDictDisease(status, value, content, note)

        # # 添加辨证
        status = 1
        # symptoms = ",".join(symptom_list)
        symptoms = ""
        sqltool.insertScySyndrome(status, disease_name, syndrome_name, symptoms)

        # 添加方案
        status = 1
        component = ""
        function = ""
        contrain = ""
        attention = ""
        sqltool.insertSchChengyao(status, disease_name, syndrome_name, name, component, function, contrain, attention)

# TODO：这里有问题
def process_sch_sytech(table):
    # 获取疾病
    diseases = sqltool.getDictDisease()
    # 获取辨证
    syndromes = sqltool.getDictSyndrome()
    # 获取技术
    sytechs = sqltool.getDictSytech()
    # 读取数据
    sytech_list = []
    for i in range(2, table.max_row):
        sytech_name = table[i][0].value
        disease_name = table[i][1].value
        syndrome_name = table[i][2].value
        symptoms = table[i][3].value
        detail = table[i][4].value
        operation = table[i][5].value
        if sytech_name is not None and len(sytech_name)>0 \
                and disease_name is not None and len(disease_name)>0 \
                and syndrome_name is not None and len(syndrome_name)>0 \
                and symptoms is not None and len(symptoms)>0:
            sytech_id = sytechs.get(sytech_name.strip())
            disease_id = diseases.get(disease_name.strip())
            syndrome_id = syndromes.get(str(disease_id) + "_" + syndrome_name.strip())
            print(disease_id, syndrome_id, sytech_id)
            if sytech_id is not None \
                    and disease_id is not None \
                    and syndrome_id is not None :
                sytech = [disease_id, syndrome_id, sytech_id, detail, operation]
                sytech_list.append(sytech)
    # 写入数据
    id = 0
    for item in sytech_list:
        id += 1
        status = 1
        disease_id = item[0]
        syndrome_id = item[1]
        sytech_id = item[2]
        detail = item[3]
        operation = item[4]
        sqltool.insertSchSytech(id, status, disease_id, syndrome_id, sytech_id, detail, operation)



if __name__ == "__main__":

    sqltool.open()

    # 加载excel
    path = '/Users/huhu/糊糊工作/脉之语/工作相关/项目相关/一真堂/资料文档/规范文档/儿科中成药指导和临床诊疗指南.xlsx'
    data = openpyxl.load_workbook(path)

    # 处理中药
    table = data["经方"]
    process_zhongyao(table)

    # 处理成药
    table = data["中成药"]
    process_chengyao(table)

    # 选择sheet（处理妇科数据）
    # table = data["中成药"]
    # process_sch_chengyao(table)

    sqltool.close()
