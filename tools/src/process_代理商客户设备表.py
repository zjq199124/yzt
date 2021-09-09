

import re
import openpyxl

from src.MysqlTool import MysqlTool


sqltool = MysqlTool()


def process_agency(table):
    for i in range(2, table.max_row+1):
        id = table[i][0].value
        name = table[i][1].value.strip()
        address = table[i][2].value.strip()
        contacts = table[i][3].value.strip()
        descrip = table[i][4].value
        status = 1
        if id is not None:
            sqltool.insertMsAgency(id, status, name, address, contacts, descrip)


def process_customer(table):
    for i in range(2, table.max_row+1):
        id = table[i][0].value
        anency_name = table[i][1].value.strip()
        name = table[i][2].value.strip().replace(" ", "")
        address = table[i][3].value.strip().replace(" ", "")
        contacts = table[i][4].value.strip().replace(" ", "")
        descrip = ""
        status = 1
        wxswitch = 1
        if id is not None:
            id = None
            sqltool.insertMsCustomer(anency_name, id, status, wxswitch, name, address, contacts, descrip)


def process_model(table):
    for i in range(2, table.max_row+1):
        id = table[i][0].value.strip()
        name = table[i][1].value.strip()
        mfrs = table[i][2].value.strip()
        descrip = table[i][3].value.strip()
        status = 1
        type = 1
        if id is not None:
            sqltool.insertTeModel(id, status, type, name, mfrs, descrip)


def process_equip(table):
    for i in range(2, table.max_row+1):
        id = table[i][0].value
        agency_name = table[i][1].value.strip()
        customer_name = table[i][2].value.strip()
        type_name = table[i][3].value.strip()
        model_name = table[i][4].value.strip()
        code = str(table[i][5].value).strip()
        buy_time = table[i][6].value
        warranty = table[i][7].value
        status = 1
        if id is not None:
            sqltool.insertTeEquip(agency_name, customer_name, type_name, model_name, id, status, code, warranty, buy_time)


if __name__ == "__main__":

    sqltool.open()

    # 加载excel
    path = '/Users/huhu/糊糊工作/脉之语/工作相关/项目相关/一真堂/资料文档/代理商客户设备表.xlsx'
    data = openpyxl.load_workbook(path)

    # 选择设备型号
    # table = data["型号"]
    # process_model(table)

    # 处理代理商
    # table = data["代理商"]
    # process_agency(table)

    # 处理客户
    # table = data["客户"]
    # process_customer(table)

    # 处理设备
    table = data["设备"]
    process_equip(table)

    sqltool.close()
